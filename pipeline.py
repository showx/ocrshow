#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""PP-StructureV3 表格识别 + 可选 Qwen3-VL。版式解析由 sheets 模块拼接。"""
from __future__ import annotations

import argparse
import base64
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import requests

from appconfig import IMAGE_EXTS, load_app_config, resolve_under_root
from ocrutil import (
    HTMLTableParser,
    extract_html_tables,
    extract_json_object,
    json_from_result,
    load_saved_structure_json,
    markdown_from_result,
    parse_date_from_stem,
    save_flatten,
    unique_keep_order,
)
import sheets

ROOT = Path(__file__).resolve().parent
VL_CHUNK_SIZE = 10


def reconstruct_from_ocr_json(payload: Any, sheet_type: str | None = None) -> list[dict[str, Any]]:
    return sheets.reconstruct(payload, sheet_type)


def records_from_saved_dir(
    image_path: Path,
    out_dir: Path,
    sheet_type: str | None = None,
) -> dict[str, Any] | None:
    image_out = out_dir / image_path.stem
    payload = load_saved_structure_json(image_out, image_path.stem)
    if payload is None:
        return None
    md_path = image_out / f"{image_path.stem}.md"
    markdown = md_path.read_text(encoding="utf-8") if md_path.exists() else ""
    records = reconstruct_from_ocr_json(payload, sheet_type)
    sheets.refine(image_path, records)
    forced = sheets.normalize_sheet_type(sheet_type)
    detected = (records[0].get("sheet_type") if records else None) or "generic"
    return {
        "image": image_path.name,
        "date": parse_date_from_stem(image_path.stem),
        "sheet_type": forced or detected,
        "markdown": markdown,
        "tables": [],
        "ocr_records": records,
        "json_payloads": [payload],
    }


def run_ppstructure(
    image_paths: list[Path],
    out_dir: Path,
    device: str,
    sheet_type: str | None = None,
) -> list[dict[str, Any]]:
    from paddleocr import PPStructureV3

    pipeline = PPStructureV3(
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
        use_formula_recognition=False,
        use_seal_recognition=False,
        use_chart_recognition=False,
        use_table_recognition=True,
        device=device,
    )

    results: list[dict[str, Any]] = []
    for image_path in image_paths:
        print(f"[PP-StructureV3] 识别 {image_path.name} ...", flush=True)
        outputs = pipeline.predict(input=str(image_path))
        image_out = out_dir / image_path.stem
        image_out.mkdir(parents=True, exist_ok=True)

        markdown_parts: list[str] = []
        html_tables: list[str] = []
        json_payloads: list[Any] = []
        for res in outputs:
            try:
                res.save_to_json(save_path=str(image_out))
            except Exception as exc:
                print(f"  保存 JSON 失败: {exc}", flush=True)
            try:
                res.save_to_markdown(save_path=str(image_out))
            except Exception as exc:
                print(f"  保存 Markdown 失败: {exc}", flush=True)
            markdown_parts.append(markdown_from_result(res))
            payload = json_from_result(res)
            json_payloads.append(payload)
            html_tables.extend(extract_html_tables(payload))

        tables: list[list[list[str]]] = []
        for html in unique_keep_order(html_tables):
            parser = HTMLTableParser()
            parser.feed(html)
            tables.extend(parser.tables)

        records: list[dict[str, Any]] = []
        for payload in json_payloads:
            records.extend(reconstruct_from_ocr_json(payload, sheet_type))
        if not records:
            records = sheets.fallback_from_tables(tables, sheet_type)

        forced = sheets.normalize_sheet_type(sheet_type)
        if forced:
            for rec in records:
                rec["sheet_type"] = forced
        sheets.refine(image_path, records)
        detected = (records[0].get("sheet_type") if records else None) or "generic"
        item = {
            "image": image_path.name,
            "date": parse_date_from_stem(image_path.stem),
            "sheet_type": forced or detected,
            "markdown": "\n\n".join(p for p in markdown_parts if p).strip(),
            "tables": tables,
            "ocr_records": records,
            "json_payloads": json_payloads,
        }
        save_flatten(image_out, item)
        print(f"  OCR 展平行数: {len(records)}", flush=True)
        results.append(item)
    return results


def _sheet_mod(ocr_item: dict[str, Any]):
    return sheets.by_id(str(ocr_item.get("sheet_type") or "")) or sheets.by_id("generic")


def call_qwen3_vl(
    image_path: Path,
    ocr_item: dict[str, Any],
    model: str,
    host: str,
    timeout: int,
) -> dict[str, Any]:
    records = ocr_item.get("ocr_records") or []
    print(f"[Qwen3-VL] 纠错整理 {image_path.name} ...", flush=True)
    if len(records) <= VL_CHUNK_SIZE:
        return call_qwen3_vl_once(image_path, ocr_item, records, model, host, timeout)
    merged: dict[str, Any] = {
        "image": image_path.name,
        "date": ocr_item.get("date"),
        "records": [],
        "corrections": [],
        "notes": [],
        "_raw_chunks": [],
    }
    for start in range(0, len(records), VL_CHUNK_SIZE):
        chunk = records[start : start + VL_CHUNK_SIZE]
        print(f"  分段 {start + 1}-{start + len(chunk)} / {len(records)}", flush=True)
        try:
            data = call_qwen3_vl_once(image_path, ocr_item, chunk, model, host, timeout)
        except Exception as exc:
            print(f"    本段失败，保留 OCR: {exc}", flush=True)
            data = {"records": chunk}
        merged["records"].extend(data.get("records") or [])
        merged["corrections"].extend(data.get("corrections") or [])
        if data.get("summary") and not merged.get("summary"):
            merged["summary"] = data.get("summary")
        if data.get("fields") and not merged.get("fields"):
            merged["fields"] = data.get("fields")
        if data.get("notes"):
            merged["notes"].append(str(data.get("notes")))
        if data.get("_raw"):
            merged["_raw_chunks"].append(data["_raw"])
    if merged["notes"]:
        merged["notes"] = "；".join(merged["notes"])
    else:
        merged["notes"] = None
    return merged


def call_qwen3_vl_once(
    image_path: Path,
    ocr_item: dict[str, Any],
    chunk_records: list[dict[str, Any]],
    model: str,
    host: str,
    timeout: int,
) -> dict[str, Any]:
    mod = _sheet_mod(ocr_item)
    ocr_text = mod.compact_ocr(chunk_records, ocr_item.get("markdown") or "")
    total = len(ocr_item.get("ocr_records") or [])
    if total > len(chunk_records):
        ocr_text = "本段只含表格中的一部分行，请只整理下面这些行。\n" + ocr_text
    prompt = mod.vl_prompt(image_path.name, str(ocr_item.get("date") or ""), ocr_text)
    image_b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
    payload = {
        "model": model,
        "stream": False,
        "format": "json",
        "think": False,
        "messages": [
            {
                "role": "user",
                "content": prompt,
                "images": [image_b64],
            }
        ],
        "options": {
            "temperature": 0.1,
            "num_ctx": 32768,
            "num_predict": 8192,
        },
    }
    url = host.rstrip("/") + "/api/chat"
    resp = requests.post(url, json=payload, timeout=timeout)
    resp.raise_for_status()
    body = resp.json()
    msg = body.get("message") or {}
    content = (msg.get("content") or "").strip() or (msg.get("thinking") or "").strip()
    print(
        f"  Qwen3-VL 完成 eval={body.get('eval_count')} content={len(msg.get('content') or '')} thinking={len(msg.get('thinking') or '')}",
        flush=True,
    )
    parsed = extract_json_object(content)
    parsed["_raw"] = content
    return parsed


def _row_key(rec: dict[str, Any]) -> int | None:
    val = rec.get("excel_row")
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def merge_vl_records(ocr_records: list[dict[str, Any]], vl_data: dict[str, Any]) -> list[dict[str, Any]]:
    vl_records = vl_data.get("records") if isinstance(vl_data, dict) else None
    if not isinstance(vl_records, list) or not vl_records:
        return ocr_records

    sheet = ocr_records[0].get("sheet_type") if ocr_records else None
    mod = sheets.by_id(str(sheet or "")) or sheets.by_id("generic")
    by_row: dict[int, dict[str, Any]] = {}
    for rec in vl_records:
        if not isinstance(rec, dict):
            continue
        key = _row_key(rec)
        if key is None:
            continue
        if str(rec.get("app_name") or rec.get("name") or rec.get("game_name") or "").strip():
            by_row.setdefault(key, rec)
    if not by_row:
        return ocr_records

    merged: list[dict[str, Any]] = []
    hit = 0
    for i, ocr in enumerate(ocr_records, start=1):
        key = _row_key(ocr)
        vl = by_row.get(key) if key is not None else None
        converted = mod.convert_vl(vl, i) if vl and mod is not None else None
        if converted is None:
            merged.append(ocr)
            continue
        hit += 1
        converted["app_name"] = ocr.get("app_name") or converted.get("app_name")
        if mod is not None:
            mod.merge_hit(ocr, converted)
        converted["excel_row"] = ocr.get("excel_row")
        converted["rank"] = ocr.get("rank") or i
        if sheet:
            converted["sheet_type"] = sheet
        merged.append(converted)
    print(f"  Qwen3-VL 覆盖 {hit}/{len(ocr_records)} 行", flush=True)
    return merged or ocr_records


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(
    path: Path,
    grouped: dict[str, list[dict[str, Any]]],
    corrections: list[dict[str, Any]],
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for mod in sheets.iter_output_modules():
        rows = grouped.get(mod.id) or []
        if first:
            ws = wb.active
            ws.title = (mod.output_stem or mod.id)[:31]
            first = False
        else:
            ws = wb.create_sheet((mod.output_stem or mod.id)[:31])
        ws.append(list(mod.csv_fields))
        for row in rows:
            ws.append([row.get(h) for h in mod.csv_fields])
    ws2 = wb.create_sheet("corrections")
    ws2.append(["image", "from", "to", "reason"])
    for item in corrections:
        ws2.append([item.get("image"), item.get("from"), item.get("to"), item.get("reason")])
    wb.save(path)


def collect_images(explicit: list[str] | None, image_dir: Path | None = None) -> list[Path]:
    if explicit:
        paths = [Path(p) if Path(p).is_absolute() else ROOT / p for p in explicit]
    else:
        folder = image_dir or (ROOT / "samples")
        if not folder.is_dir():
            raise FileNotFoundError(
                f"未找到截图目录 {folder}。请把图片放到 samples/，或用参数指定路径。"
            )
        paths = sorted(
            p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTS
        )
    existing = [p for p in paths if p.exists()]
    missing = [p for p in paths if not p.exists()]
    for p in missing:
        print(f"跳过不存在的文件: {p}", flush=True)
    if not existing:
        raise FileNotFoundError(
            f"未找到待识别图片。请把截图放到 {image_dir or ROOT / 'samples'}，或在命令行传入路径。"
        )
    return existing


def pick_device(prefer: str) -> str:
    if prefer and prefer != "auto":
        return prefer
    try:
        import paddle

        if paddle.device.cuda.device_count() > 0:
            return "gpu"
    except Exception:
        pass
    return "cpu"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="PP-StructureV3 + 可插拔版式模块")
    parser.add_argument("images", nargs="*", help="图片路径；省略时读取配置中的截图目录（默认 samples/）")
    parser.add_argument("--config", default="", help="额外 TOML 配置，覆盖 config.toml")
    parser.add_argument("--images-dir", default="", help="截图目录，默认读取配置")
    parser.add_argument("--out", default="", help="输出目录，默认读取配置")
    parser.add_argument("--device", default="", help="PP-StructureV3 设备: auto/gpu/cpu")
    parser.add_argument("--vl-model", default="", help="Ollama 中的 Qwen3-VL 模型名")
    parser.add_argument("--vl-host", default="", help="Ollama 地址")
    parser.add_argument("--skip-vl", action="store_true", help="只跑 OCR，不调用 Qwen3-VL")
    parser.add_argument("--with-vl", action="store_true", help="强制调用 Qwen3-VL（覆盖配置）")
    parser.add_argument("--from-ocr", action="store_true", help="复用已有 PP-StructureV3 JSON，不再重新识别")
    parser.add_argument("--from-vl", action="store_true", help="复用已有 qwen3vl.json，不再调用模型")
    parser.add_argument("--timeout", type=int, default=0, help="Qwen3-VL 请求超时秒数，0 表示用配置")
    parser.add_argument("--sheet-type", default="auto", help="强制版式：auto 或已安装模块 id")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    extra = Path(args.config) if args.config else None
    cfg = load_app_config(extra)
    paths = cfg.get("paths") or {}
    vl = cfg.get("vl") or {}
    sheets.load_modules()

    out_dir = resolve_under_root(args.out or paths.get("output"), "output")
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.with_vl:
        args.skip_vl = False
    elif not args.skip_vl:
        args.skip_vl = bool(vl.get("skip"))
    args.vl_model = args.vl_model or str(vl.get("model") or "qwen3-vl:8b")
    args.vl_host = args.vl_host or str(vl.get("host") or "http://127.0.0.1:11434")
    if not args.timeout:
        try:
            args.timeout = int(vl.get("timeout") or 600)
        except (TypeError, ValueError):
            args.timeout = 600

    forced_sheet = sheets.normalize_sheet_type(args.sheet_type)
    images_dir = resolve_under_root(args.images_dir or paths.get("images"), "samples")
    images = collect_images(args.images, images_dir)
    device = pick_device(args.device or str(cfg.get("device") or "auto"))
    print(f"设备: {device}", flush=True)
    print(f"类别: {forced_sheet or 'auto'}", flush=True)
    print(f"模块: {', '.join(m.id for m in sheets.load_modules())}", flush=True)
    print(f"图片: {', '.join(p.name for p in images)}", flush=True)

    if args.from_ocr:
        ocr_items = []
        for image_path in images:
            item = records_from_saved_dir(image_path, out_dir, forced_sheet)
            if item is None:
                raise FileNotFoundError(f"未找到 {image_path.stem} 的 OCR 结果，请先跑识别")
            save_flatten(out_dir / image_path.stem, item)
            print(f"[复用 OCR] {image_path.name} 展平行数: {len(item['ocr_records'])}", flush=True)
            ocr_items.append(item)
    else:
        ocr_items = run_ppstructure(images, out_dir, device, forced_sheet)

    all_rows: list[dict[str, Any]] = []
    all_corrections: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []

    for image_path, ocr_item in zip(images, ocr_items):
        vl_data: dict[str, Any] | None = None
        ocr_n = len(ocr_item.get("ocr_records") or [])
        vl_path = out_dir / image_path.stem / "qwen3vl.json"
        if args.from_vl and vl_path.exists():
            vl_data = json.loads(vl_path.read_text(encoding="utf-8"))
            print(f"  复用 Qwen3-VL {vl_path.name}", flush=True)
        elif not args.skip_vl and ocr_n > 100:
            print(f"  记录过多({ocr_n})，跳过 Qwen3-VL，保留 OCR", flush=True)
        elif not args.skip_vl:
            try:
                vl_data = call_qwen3_vl(
                    image_path,
                    ocr_item,
                    model=args.vl_model,
                    host=args.vl_host,
                    timeout=args.timeout,
                )
                (out_dir / image_path.stem / "qwen3vl.json").write_text(
                    json.dumps(vl_data, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
            except Exception as exc:
                print(f"  Qwen3-VL 失败，保留 OCR 结果: {exc}", flush=True)

        records = merge_vl_records(ocr_item.get("ocr_records") or [], vl_data or {})
        sheet_type = forced_sheet or ocr_item.get("sheet_type") or "generic"
        for rec in records:
            rec["image"] = ocr_item["image"]
            rec["date"] = ocr_item.get("date")
            rec["sheet_type"] = sheet_type if forced_sheet else rec.get("sheet_type") or sheet_type
        all_rows.extend(records)

        corrections = []
        if vl_data:
            for item in vl_data.get("corrections") or []:
                if isinstance(item, dict):
                    item = dict(item)
                    item["image"] = ocr_item["image"]
                    corrections.append(item)
        all_corrections.extend(corrections)

        mod = sheets.by_id(str(sheet_type)) or sheets.by_id("generic")
        default_fields = dict(mod.fields) if mod else {}
        summaries.append(
            {
                "image": ocr_item["image"],
                "date": ocr_item.get("date"),
                "sheet_type": sheet_type,
                "ocr_count": len(ocr_item.get("ocr_records") or []),
                "final_count": len(records),
                "summary": (vl_data or {}).get("summary"),
                "notes": (vl_data or {}).get("notes"),
                "fields": (vl_data or {}).get("fields") or default_fields,
            }
        )
        print(f"  最终记录数: {len(records)}", flush=True)

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for rec in all_rows:
        grouped[str(rec.get("sheet_type") or "generic")].append(rec)
    for sid, rows in grouped.items():
        rows.sort(
            key=lambda r: (
                str(r.get("date") or ""),
                int(r.get("col_group") or 0),
                int(r.get("excel_row") or 0),
                str(r.get("image") or ""),
                int(r.get("rank") or 0),
            )
        )
        for i, rec in enumerate(rows, start=1):
            rec["global_rank"] = i

    totals = {f"{m.id}_total": len(grouped.get(m.id) or []) for m in sheets.iter_output_modules()}
    (out_dir / "summary.json").write_text(
        json.dumps({"images": summaries, **totals, "total": len(all_rows)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (out_dir / "records.json").write_text(json.dumps(all_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    for mod in sheets.iter_output_modules():
        rows = grouped.get(mod.id) or []
        stem = mod.output_stem or mod.id
        (out_dir / f"{stem}.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        write_csv(out_dir / f"{stem}.csv", rows, list(mod.csv_fields))
    try:
        write_xlsx(out_dir / "records.xlsx", dict(grouped), all_corrections)
    except Exception as exc:
        print(f"写入 xlsx 失败: {exc}", flush=True)

    result_payload = {
        "ok": True,
        "category": forced_sheet or "auto",
        "records": all_rows,
        "summary": {"images": summaries, **totals, "total": len(all_rows)},
    }
    (out_dir / "result.json").write_text(
        json.dumps(result_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    parts = ", ".join(f"{m.name} {len(grouped.get(m.id) or [])} 条" for m in sheets.iter_output_modules())
    print(f"\n完成。{parts}，输出目录: {out_dir}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
